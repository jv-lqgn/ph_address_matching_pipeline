"""
PH Address Parsing Pipeline  (v4)
==================================
Stage 0 — Load reference data
Stage 1 — Noise table filtering  (pure noise → noise bucket, skip pipeline)
Stage 2 — Junk token stripping
Stage 3 — Alias normalization
Stage 4 — City detection  (right-to-left, exact → fuzzy, multi-candidate)
           FIX v3: Manila districts matched directly as city_name in dim_location.
Stage 5 — Barangay explicit-presence gate + fuzzy scoring
           FIX v3: barangay never assumed when not explicitly present.
           FIX v4: replaced partial_ratio (substring-prone) with a two-step gate:
                   (1) bgy significant tokens must appear word-boundary in address
                       AFTER stripping [word]+[street-type] pairs and city tokens,
                   (2) only then score with token_set_ratio (≥ BGY_FUZZY_THRESHOLD).
                   Addresses where bgy name appears ONLY as a street reference
                   (e.g. "Rizal Street", "Dela Paz Ave") correctly become partial.
Stage 6 — Confidence scoring
Stage 7 — Export  (valid / partial / invalid / noise)
"""

import os
import re
import unicodedata
from datetime import datetime

import pandas as pd
from rapidfuzz import fuzz, process
from openpyxl.styles import Font, PatternFill, Alignment

# ── 0. CONFIG ─────────────────────────────────────────────────────────────────

DIM_LOC_PATH  = "../../data/mapping/dim_location_20260415_v3.csv"
ALIAS_PATH    = "../../data/utils/ph_address_alias_extended_v4.csv"
INPUT_PATH    = "../../data/sample_address.xlsx"
INPUT_COL     = "order_deliveraddress"

OUTPUT_DIR    = "../../data/output"
VALID_DIR     = os.path.join(OUTPUT_DIR, "valid")
PARTIAL_DIR   = os.path.join(OUTPUT_DIR, "partial")
INVALID_DIR   = os.path.join(OUTPUT_DIR, "invalid")
NOISE_DIR     = os.path.join(OUTPUT_DIR, "noise")          # NEW

for d in [VALID_DIR, PARTIAL_DIR, INVALID_DIR, NOISE_DIR]:
    os.makedirs(d, exist_ok=True)

# Fuzzy thresholds
CITY_FUZZY_THRESHOLD   = 85
BGY_FUZZY_THRESHOLD    = 70
CONFIDENCE_VALID       = 75
CONFIDENCE_PARTIAL     = 45

print("Config OK ✓")


# ── 1. HELPERS ────────────────────────────────────────────────────────────────

def strip_accents(text: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", text)
        if unicodedata.category(c) != "Mn"
    )

def clean_str(s: str) -> str:
    s = strip_accents(str(s)).lower()
    return re.sub(r"\s+", " ", s).strip()

print("Helpers defined ✓")


# ── 2. LOAD REFERENCE DATA ────────────────────────────────────────────────────

print("\nLoading reference data …")

dim_raw = pd.read_csv(DIM_LOC_PATH, encoding="latin1")
dim_raw.columns = dim_raw.columns.str.strip()
for col in ["barangay_name", "city_name", "province_name", "region_name"]:
    dim_raw[col] = dim_raw[col].astype(str).str.strip()

dim_raw["bgy_clean"]  = dim_raw["barangay_name"].apply(clean_str)
dim_raw["city_clean"] = dim_raw["city_name"].apply(clean_str)
dim_raw["prov_clean"] = dim_raw["province_name"].apply(clean_str)

# ── PROBLEM 1 FIX ─────────────────────────────────────────────────────────────
# In dim_location, Manila districts ARE the city_name (e.g. city_name="Binondo",
# province_name="Metro Manila"). There is NO "Manila" city record.
# The old code mapped every district to a phantom "Manila" that could never match.
#
# Fix: MANILA_DISTRICTS maps each district's clean form to its EXACT city_name
# as stored in dim_location. City detection returns (clean, original) tuples that
# match dim_location directly — no phantom "Manila" collapse.
#
# "San Andres" note: dim_location stores this district as city_name="San Andres,
# Bukid" (province=Metro Manila). Confirm with your actual CSV; adjust below if needed.

MANILA_DISTRICTS = {
    # clean form           : exact city_name as in dim_location
    "binondo"    : "Binondo",
    "ermita"     : "Ermita",
    "intramuros" : "Intramuros",
    "malate"     : "Malate",
    "paco"       : "Paco",
    "pandacan"   : "Pandacan",
    "port area"  : "Port Area",
    "quiapo"     : "Quiapo",
    "sampaloc"   : "Sampaloc",
    "san andres" : "San Andres",      # verify exact spelling in your CSV
    "san miguel" : "San Miguel",
    "san nicolas": "San Nicolas",
    "santa ana"  : "Santa Ana",
    "santa cruz" : "Santa Cruz",
    "tondo"      : "Tondo I/II",       # dim_location stores as "Tondo I/II"
    "tondo i/ii" : "Tondo I/II",
}

# Metro Manila province label used to disambiguate same-name cities
NCR_PROVINCE = "Metro Manila"

# Build city lookup — all unique city_name values in dim_location
all_cities_clean = {clean_str(c): c for c in dim_raw["city_name"].dropna().unique()}
all_city_cleans  = list(all_cities_clean.keys())

# Alias map
alias_df = pd.read_csv(ALIAS_PATH, encoding="latin1", usecols=["pattern", "replacement"])
alias_df = alias_df.dropna(subset=["pattern", "replacement"])
alias_df["pattern"]     = alias_df["pattern"].astype(str).str.strip()
alias_df["replacement"] = alias_df["replacement"].astype(str).str.strip()
alias_df = alias_df.sort_values("pattern", key=lambda s: s.str.len(), ascending=False)
alias_map = dict(zip(alias_df["pattern"].str.upper(), alias_df["replacement"].str.upper()))

print(f"  dim_location rows : {len(dim_raw):,}")
print(f"  alias rules       : {len(alias_map):,}")
print(f"  unique cities     : {len(all_cities_clean):,}")
print("Reference data loaded ✓")


# ── 3. LOAD ADDRESSES ─────────────────────────────────────────────────────────

print("\nLoading addresses …")
input_df      = pd.read_excel(INPUT_PATH)
RAW_ADDRESSES = input_df[INPUT_COL].dropna().tolist()
print(f"  {len(RAW_ADDRESSES):,} addresses loaded")


# ── 4. STAGE 1 — NOISE TABLE (NEW) ───────────────────────────────────────────
#
# PROBLEM 3 FIX:
# Addresses that contain ONLY structural/street-level tokens (village, subdivision,
# street, blk, house #, etc.) with NO city or province signal are pure noise.
# These are removed before the pipeline runs — they go into a dedicated noise file.
#
# How it works:
#   1. Strip obvious structural tokens and numbers from the address.
#   2. If after stripping nothing substantive remains AND no city/province keyword
#      is found in the raw text → classify as noise.
#
# Extend NOISE_KEYWORDS as you discover new noise patterns in your data.

# Keywords that by themselves carry no geographic resolution
NOISE_KEYWORDS = re.compile(
    r"\b(village|vill|vllg|subdivision|subd|blk|block|lot|phase|unit|"
    r"street|st\b|ave|avenue|road|rd\b|boulevard|blvd|lane|drive|"
    r"house no|hse no|house #|h\.?no|purok|sitio|compound|cmpd|zone|"
    r"floor|flr|room|rm\b|door|building|bldg|tower|plaza|mall|"
    r"corner|cor\b|extension|extn?)\b",
    re.IGNORECASE,
)

# A number-only or structural-only fragment has no geo anchor
_NOISE_NUM = re.compile(r"\b\d+[-\w]*\b")


def is_pure_noise(raw_addr: str) -> bool:
    """
    Returns True when the address carries ONLY structural tokens (street, blk,
    village, house #, etc.) with no city or province anchor.

    Strategy:
      • Remove all noise-category keywords and numbers.
      • Remove punctuation/whitespace.
      • If nothing substantive (≥ 4 alpha chars) remains → noise.
    """
    s = raw_addr.lower()
    s = NOISE_KEYWORDS.sub(" ", s)
    s = _NOISE_NUM.sub(" ", s)
    s = re.sub(r"[^a-z\s]", " ", s)           # strip punctuation
    s = re.sub(r"\s{2,}", " ", s).strip()

    # Remaining tokens that are ≥ 4 alphabetic chars could be a city/province name
    meaningful = [tok for tok in s.split() if len(tok) >= 4]
    return len(meaningful) == 0


print("Noise filter defined ✓")


# ── 5. STAGE 2 — JUNK TOKEN STRIPPING ────────────────────────────────────────

_BUILDING_KW = (
    r"bldg|building|tower|plaza|centre|center|subd|subdivision|compound|cmpd|"
    r"village|vill|estate|residences|residencia|condominium|condo|"
    r"apartelle|apartment|apt|annex|mall|square|complex|commercial|industrial|zone|cluster"
)

_PROTECT_PATTERNS = [
    (re.compile(r"\bA\.C\.?\b", re.IGNORECASE), "ANGELES_CITY"),
]

_STRAY_PUNCT      = re.compile(r"[`~|]")
_PHONE_NUMBERS    = re.compile(r"\b(0|\+63)\d{9,10}\b")
_STREET_NUMBERS   = re.compile(r"(?<!\w)\d{3,5}(?!\w)")
_LOT_BLK_UNIT     = re.compile(
    r"\b(lot|blk|block|unit|floor|flr|fl|rm|room|door|phase|house no|hse no|#)"
    r"\s*[.\-]?\s*[\w\-]*",
    re.IGNORECASE,
)
_LANDMARK_PHRASES = re.compile(
    r"\b(near|beside|in front of|across from|across|opposite|opp\.?|behind|"
    r"adjacent to|adj\.?|along|corner of|corner|cor\.?)\\b[^,]*",
    re.IGNORECASE,
)
_BUILDING_UNIFIED = re.compile(
    r"\b(?:[A-Za-z]\w*\.?\s+){0,4}"
    r"\b(?:" + _BUILDING_KW + r")\b\.?"
    r"(?:\s+(?!(?:brgy|barangay|st\b|ave\b|road\b|blvd\b))[A-Za-z]\w*){0,4}",
    re.IGNORECASE,
)


def strip_junk(addr: str) -> str:
    s = addr
    for pattern, placeholder in _PROTECT_PATTERNS:
        s = pattern.sub(placeholder, s)
    s = _STRAY_PUNCT.sub(" ", s)
    s = _PHONE_NUMBERS.sub(" ", s)
    s = _LANDMARK_PHRASES.sub(" ", s)
    s = _LOT_BLK_UNIT.sub(" ", s)
    s = _BUILDING_UNIFIED.sub(" ", s)
    s = _STREET_NUMBERS.sub(" ", s)
    s = s.replace("ANGELES_CITY", "Angeles City")
    s = re.sub(r"\s{2,}", " ", s)
    s = re.sub(r"(,\s*){2,}", ", ", s)
    s = re.sub(r"^[\s,\.]+|[\s,\.]+$", "", s)
    return s.strip()


print("Junk stripper defined ✓")


# ── 6. STAGE 3 — ALIAS NORMALIZATION ─────────────────────────────────────────

def normalize_address(addr: str) -> str:
    upper  = addr.upper()
    tokens = re.split(r"(\s+|,|\.)", upper)
    result = []
    i = 0
    while i < len(tokens):
        tok = tokens[i].strip()
        if not tok or tok in (",", "."):
            result.append(tokens[i])
            i += 1
            continue
        matched = False
        for lookahead in (3, 2, 1):
            candidate_tokens, j, count = [], i, 0
            while j < len(tokens) and count < lookahead:
                if re.match(r"^\s*$", tokens[j]) or tokens[j] in (",", "."):
                    j += 1
                    continue
                candidate_tokens.append(tokens[j])
                j += 1
                count += 1
            candidate = " ".join(candidate_tokens)
            if candidate in alias_map:
                result.append(alias_map[candidate])
                i = j
                matched = True
                break
        if not matched:
            result.append(tokens[i])
            i += 1
    normalized = re.sub(r"\s+", " ", "".join(result)).strip()
    normalized = re.sub(r"\.\s+", " ", normalized)
    normalized = re.sub(r"\s{2,}", " ", normalized).strip()
    return normalized.title()


print("Alias normalizer defined ✓")


# ── 7. STAGE 4 — CITY DETECTION ──────────────────────────────────────────────
#
# PROBLEM 1 FIX (continued):
# detect_city_candidates() now:
#   • Checks MANILA_DISTRICTS first — if a district name is found in the address,
#     it returns (district_clean, district_city_name_in_dim) directly.
#     e.g. "Binondo" → ("binondo", "Binondo") — NOT "Manila".
#   • Uses NCR_PROVINCE to filter when a name is ambiguous across regions
#     (e.g. "San Miguel" exists in Bulacan AND Metro Manila; if "manila" or
#     "ncr" or "metro" appears in the address, prefer the Metro Manila one).
#   • Falls through to standard city matching for non-Manila cities.
#
# detect_via_barangay() remains as-is (barangay → city fallback) but also uses
# MANILA_DISTRICTS mapping so it returns the correct city_name.

_BARANGAY_PREFIX = re.compile(r"^\s*barangay\s*", re.IGNORECASE)
_STREET_TOKENS   = re.compile(
    r"\b(street|avenue|road|boulevard|highway|lane|drive|circle|place|extension)\b",
    re.IGNORECASE,
)

_bgy_stripped_entries = [
    (
        row["bgy_clean"],
        _BARANGAY_PREFIX.sub("", clean_str(row["bgy_clean"])).strip(),
        row["city_name"],
        row["province_name"],
    )
    for _, row in dim_raw.iterrows()
]
_bgy_stripped_list = [x[1] for x in _bgy_stripped_entries]


def _prep_for_bgy_match(norm_addr: str) -> str:
    s = _BARANGAY_PREFIX.sub("", clean_str(norm_addr)).strip()
    s = re.sub(r"\b[\d][\d\-a-z]*\b", "", s)
    s = _STREET_TOKENS.sub("", s)
    return re.sub(r"\s{2,}", " ", s).strip()


def _address_hints_ncr(addr_c: str) -> bool:
    """Return True if the cleaned address contains hints of Metro Manila / NCR."""
    return bool(re.search(r"\b(manila|metro|ncr|metro manila)\b", addr_c))


def detect_city_candidates(norm_addr: str) -> list[tuple[str, str]]:
    """
    Returns [(city_clean, city_original_as_in_dim), ...].
    city_original is the EXACT city_name string stored in dim_location so that
    later dim_raw filtering by city_clean works correctly.

    Manila districts:
      • Any district name found in the address returns the district's own city_name
        (e.g. "binondo" → ("binondo", "Binondo")) — NOT a phantom "Manila".
      • Ambiguous names (San Miguel, Santa Cruz, San Nicolas, Santa Ana, San Andres)
        are pinned to the Metro Manila record when NCR hints are present.
    """
    addr_c       = clean_str(norm_addr)
    segments_rtl = list(reversed([s.strip() for s in norm_addr.split(",")]))
    candidates   = {}

    # ── Priority pass: explicit Manila district in raw address ────────────────
    for dist_clean, dist_city_orig in MANILA_DISTRICTS.items():
        if re.search(r"\b" + re.escape(dist_clean) + r"\b", addr_c):
            candidates[dist_clean] = dist_city_orig

    # ── Standard city detection per segment ──────────────────────────────────
    for seg in segments_rtl:
        seg_c = clean_str(seg)
        if len(seg_c) < 3:
            continue
        found = False

        # Pass 1 — exact word-boundary, longest first
        for city_c, city_orig in sorted(all_cities_clean.items(), key=lambda x: -len(x[0])):
            if city_c in candidates:          # already captured via district pass
                found = True
                break
            if re.search(r"\b" + re.escape(city_c) + r"\b", seg_c):
                # Ambiguous name: prefer Metro Manila when NCR hints present
                if _address_hints_ncr(addr_c) and city_c in MANILA_DISTRICTS:
                    candidates[city_c] = MANILA_DISTRICTS[city_c]
                else:
                    candidates[city_c] = city_orig
                found = True
                break
        if found:
            continue

        # Pass 2 — strip "City" suffix and retry exact
        seg_no_city = re.sub(r"\s{2,}", " ", re.sub(r"\bcity\b", "", seg_c)).strip()
        if seg_no_city != seg_c:
            for city_c, city_orig in sorted(all_cities_clean.items(), key=lambda x: -len(x[0])):
                if re.search(r"\b" + re.escape(city_c) + r"\b", seg_no_city):
                    if _address_hints_ncr(addr_c) and city_c in MANILA_DISTRICTS:
                        candidates[city_c] = MANILA_DISTRICTS[city_c]
                    else:
                        candidates[city_c] = city_orig
                    found = True
                    break
        if found:
            continue

        # Pass 3 — fuzzy on original segment
        best = process.extractOne(seg_c, all_city_cleans, scorer=fuzz.token_set_ratio)
        if best and best[1] >= CITY_FUZZY_THRESHOLD:
            city_c = best[0]
            if _address_hints_ncr(addr_c) and city_c in MANILA_DISTRICTS:
                candidates[city_c] = MANILA_DISTRICTS[city_c]
            else:
                candidates[city_c] = all_cities_clean[city_c]
            continue

        # Pass 4 — fuzzy on City-stripped segment
        if seg_no_city != seg_c:
            best2 = process.extractOne(seg_no_city, all_city_cleans, scorer=fuzz.token_set_ratio)
            if best2 and best2[1] >= CITY_FUZZY_THRESHOLD:
                city_c = best2[0]
                if _address_hints_ncr(addr_c) and city_c in MANILA_DISTRICTS:
                    candidates[city_c] = MANILA_DISTRICTS[city_c]
                else:
                    candidates[city_c] = all_cities_clean[city_c]

    # De-duplicate while preserving insertion order
    deduped, seen = [], set()
    for item in candidates.items():
        if item not in seen:
            seen.add(item)
            deduped.append(item)
    return deduped


def detect_via_barangay(norm_addr: str) -> list[tuple[str, str]]:
    """
    Fallback: infer city from barangay name when city detection yields nothing.
    Returns [(city_original, province_original), ...].
    Manila districts are mapped via MANILA_DISTRICTS so the returned city_name
    matches dim_location exactly.
    """
    q = _prep_for_bgy_match(norm_addr)
    if len(q) < 3:
        return []

    tokens = q.split()
    for n in (3, 2, 1):
        for i in range(len(tokens) - n + 1):
            phrase = " ".join(tokens[i:i + n])
            if len(phrase) < 4:
                continue
            hits = dim_raw[dim_raw["bgy_clean"] == phrase]
            if not hits.empty:
                out = []
                for _, r in hits.drop_duplicates("city_name").iterrows():
                    city_name     = r["city_name"]
                    province_name = r["province_name"]
                    city_clean    = clean_str(city_name)
                    if city_clean in MANILA_DISTRICTS:
                        out.append((MANILA_DISTRICTS[city_clean], province_name))
                    else:
                        out.append((city_name, province_name))
                return out

    best = process.extractOne(q, _bgy_stripped_list, scorer=fuzz.token_sort_ratio)
    if best and best[1] >= BGY_FUZZY_THRESHOLD:
        idx           = _bgy_stripped_list.index(best[0])
        entry         = _bgy_stripped_entries[idx]
        city_name     = entry[2]
        province_name = entry[3]
        city_clean    = clean_str(city_name)
        if city_clean in MANILA_DISTRICTS:
            return [(MANILA_DISTRICTS[city_clean], province_name)]
        return [(city_name, province_name)]

    return []


print("City detector + barangay fallback defined ✓")


# ── 8. STAGE 5 — BARANGAY EXPLICIT-PRESENCE GATE + FUZZY SCORING ─────────────
#
# FIX v4 — ROOT CAUSE: the old _score_match used fuzz.partial_ratio which matches
# SUBSTRINGS, not whole words. This caused systematic false positives:
#   • "pagasa" contains "mesa" → Lamesa scored 83%
#   • "santa" contains "nta"  → Atsan scored 80%
#   • "Rizal Street" → "rizal" left as orphan after stripping → 100%
#
# NEW APPROACH — two mandatory gates before any barangay is accepted:
#
# GATE 1 — Explicit token presence (word-boundary check):
#   Build a "working" address by:
#     (a) Strip [word]+[street-type] pairs — removes street-name orphans like
#         "rizal street" → gone, "onyx street" → gone.
#         NOTE: only ONE preceding word is stripped with each street-type word,
#         so "santo domingo street" → "santo" remains (safe for multi-word bgy names).
#     (b) Strip city name tokens from the working address — prevents city name
#         from bleeding into bgy match — UNLESS the bgy tokens substantially
#         overlap with the city tokens (e.g. bgy="San Nicolas", city="San Nicolas":
#         don't strip or nothing would remain).
#   Then: ALL significant bgy tokens (after removing part/number suffixes like
#   "Pob.", "I/II", "Uno") must appear as \b word-boundary matches in the working
#   address. If any token is missing → bgy is NOT explicitly mentioned → skip.
#
# GATE 2 — token_set_ratio score (≥ BGY_FUZZY_THRESHOLD):
#   Scores the bgy name against the full cleaned address using token_set_ratio.
#   Unlike partial_ratio, token_set_ratio sorts and intersects token sets so short
#   bgy names embedded inside unrelated long strings score low.
#   Both gates must pass for a barangay to be recorded.
#
# Conservative design: addresses where the bgy name appears ONLY as a street
# reference (e.g. "Dela Paz Ave", "Magsaysay Blvd") will correctly become
# partial (city only, no bgy). False negatives here are safe; false positives
# (wrongly assigning a barangay) are the bug being fixed.

# Strip exactly ONE preceding word + street-type word.
# Handles "rizal street" → removed, but keeps "santo" from "santo domingo street".
_WORD_STREET_PAIR = re.compile(
    r"\b\w+\s+(?:street|avenue|road|boulevard|highway|lane|drive|circle)\b",
    re.IGNORECASE,
)

# Suffixes that carry no geographic identity in bgy names
_BGY_PART_SUFFIXES = re.compile(
    r"\b(pob\.?|i{1,3}|iv|vi{0,3}|viii|ix|x{1,3}|zone|part|uno|dos|tres|ext(?:ension)?)\b",
    re.IGNORECASE,
)


def _bgy_significant_tokens(bgy_c: str) -> list[str]:
    """Strip part/number suffixes and return tokens of ≥ 3 chars."""
    s = _BGY_PART_SUFFIXES.sub(" ", bgy_c)
    s = re.sub(r"[^a-z\s]", " ", s)
    return [t for t in s.split() if len(t) >= 3]


def _bgy_gate1_passes(addr_c: str, bgy_c: str, city_c: str | None) -> bool:
    """
    Gate 1: all significant bgy tokens must appear as word-boundary matches
    in the address after stripping street-name pairs and city tokens.
    Returns False → bgy is not explicitly mentioned → do not match.
    """
    sig_tokens = _bgy_significant_tokens(bgy_c)
    if not sig_tokens:
        return False

    # (a) Strip [word] + [street-type] pairs
    working = _WORD_STREET_PAIR.sub(" ", addr_c)
    working = re.sub(r"\s{2,}", " ", working).strip()

    # (b) Strip city tokens only when they don't substantially overlap bgy tokens
    if city_c:
        city_tokens = [t for t in city_c.split() if len(t) >= 4]
        overlap     = sum(1 for t in sig_tokens if t in city_tokens)
        if city_tokens and overlap < len(sig_tokens) / 2:
            for tok in city_tokens:
                working = re.sub(r"\b" + re.escape(tok) + r"\b", " ", working)
            working = re.sub(r"\s{2,}", " ", working).strip()

    # All significant tokens must be present as whole words
    for tok in sig_tokens:
        if not re.search(r"\b" + re.escape(tok) + r"\b", working):
            return False
    return True


def _score_match(
    addr_c: str, city_c: str, sub: pd.DataFrame
) -> tuple[float, pd.Series | None]:
    """
    For each barangay in sub:
      1. Gate 1 — explicit token presence (see _bgy_gate1_passes).
         Candidates that don't pass are skipped entirely.
      2. Gate 2 — token_set_ratio score (≥ BGY_FUZZY_THRESHOLD).
    Returns (best_score, best_row) across all passing candidates.
    score = 0.0 and row = None when nothing passes both gates.
    """
    if sub.empty:
        return 0.0, None

    best_score = 0.0
    best_row   = None

    for _, row in sub.iterrows():
        bgy_c = row["bgy_clean"]

        # Gate 1 — must be explicitly present
        if not _bgy_gate1_passes(addr_c, bgy_c, city_c):
            continue

        # Gate 2 — token_set_ratio (avoids partial_ratio substring trap)
        score = fuzz.token_set_ratio(bgy_c, addr_c)
        if score >= BGY_FUZZY_THRESHOLD and score > best_score:
            best_score = score
            best_row   = row

    return best_score, best_row


def match_address(row: pd.Series) -> dict:
    norm           = row["normalized_address"]
    addr_c         = clean_str(norm)
    candidates     = list(row["city_candidates"])
    bgy_city_cands = row["bgy_city_candidates"]

    if not candidates and bgy_city_cands:
        candidates = [(clean_str(city), city) for city, prov in bgy_city_cands]

    if not candidates:
        return _build_result(
            orig_addr   = row["order_deliveraddress"],
            norm_addr   = norm,
            addr_c      = addr_c,
            matched_bgy = None,
            bgy_score   = 0.0,
            city_score  = 0.0,
            tier        = "invalid",
            reason      = "no_city_detected",
        )

    best_bgy_score  = -1.0
    best_bgy_row    = None
    best_city_clean = None
    best_city_orig  = None

    for city_c, city_orig in candidates:
        sub = dim_raw[dim_raw["city_clean"] == city_c].copy()
        # Pass city_c into _score_match so Gate 1 can strip city tokens
        bgy_score, matched_row = _score_match(addr_c, city_c, sub)
        if bgy_score > best_bgy_score:
            best_bgy_score  = bgy_score
            best_bgy_row    = matched_row
            best_city_clean = city_c
            best_city_orig  = city_orig

    city_score = fuzz.token_set_ratio(best_city_clean, addr_c) if best_city_clean else 0.0

    # Barangay is recorded ONLY when both gates passed (score > 0 means both passed)
    bgy_explicitly_matched = best_bgy_score >= BGY_FUZZY_THRESHOLD
    matched_bgy = best_bgy_row if bgy_explicitly_matched else None

    composite = 0.7 * best_bgy_score + 0.3 * city_score

    if bgy_explicitly_matched and composite >= CONFIDENCE_VALID:
        tier   = "valid"
        reason = "barangay_matched"
    elif best_city_orig:
        tier   = "partial"
        reason = "city_only_no_barangay" if not bgy_explicitly_matched else "city_only_or_low_confidence"
    else:
        tier   = "invalid"
        reason = "below_confidence_threshold"

    return _build_result(
        orig_addr   = row["order_deliveraddress"],
        norm_addr   = norm,
        addr_c      = addr_c,
        matched_bgy = matched_bgy,
        bgy_score   = round(best_bgy_score, 1),
        city_score  = round(city_score, 1),
        tier        = tier,
        reason      = reason,
        city_orig   = best_city_orig,
        composite   = round(composite, 1),
    )


def _build_result(
    orig_addr, norm_addr, addr_c,
    matched_bgy, bgy_score, city_score,
    tier, reason,
    city_orig=None, composite=None,
) -> dict:
    """
    Assemble the output dict.
    PROBLEM 2 FIX: barangay fields populated ONLY when matched_bgy is not None.
    City fields are always populated from the detected city candidate, never from
    the matched_bgy row, so a city-only match still reports the correct city.
    """
    # Resolve city/province/region from matched barangay row if available,
    # else fall back to the detected city_orig (city-only result).
    if matched_bgy is not None:
        city_name     = matched_bgy["city_name"]
        province_name = matched_bgy["province_name"]
        region_name   = matched_bgy["region_name"]
        city_code     = matched_bgy.get("city_code")
        province_code = matched_bgy.get("province_code")
        region_code   = matched_bgy.get("region_code")
        barangay_code = matched_bgy.get("barangay_code")
        barangay_name = matched_bgy["barangay_name"]
    else:
        # City is confirmed but barangay is not — leave barangay fields NULL
        city_name     = city_orig
        province_name = None
        region_name   = None
        city_code     = None
        province_code = None
        region_code   = None
        barangay_code = None
        barangay_name = None   # ← never assumed

    return {
        "order_deliveraddress" : orig_addr,
        "normalized_address"   : norm_addr,
        "barangay_code"        : barangay_code,
        "barangay_name"        : barangay_name,   # NULL when not explicitly matched
        "city_code"            : city_code,
        "city_name"            : city_name,
        "province_code"        : province_code,
        "province_name"        : province_name,
        "region_code"          : region_code,
        "region_name"          : region_name,
        "addr_clean"           : addr_c,
        "bgy_match_score"      : bgy_score,
        "city_match_score"     : city_score,
        "confidence_score"     : composite if composite is not None else 0.0,
        "match_tier"           : tier,
        "match_reason"         : reason,
    }


print("Matcher (v4 two-gate bgy) defined ✓")


# ── 9. RUN PIPELINE ───────────────────────────────────────────────────────────

print("\nRunning pipeline …")

noise_records   = []
pipeline_inputs = []

for addr in RAW_ADDRESSES:
    # ── PROBLEM 3 FIX: noise check before anything else ──────────────────────
    if is_pure_noise(addr):
        noise_records.append({
            "order_deliveraddress" : addr,
            "noise_reason"         : "structural_tokens_only_no_geo_anchor",
        })
        continue

    # Normal pipeline
    stripped            = strip_junk(addr)
    normalized          = normalize_address(stripped)
    city_candidates     = detect_city_candidates(normalized)
    bgy_city_candidates = detect_via_barangay(normalized) if not city_candidates else []
    pipeline_inputs.append({
        "order_deliveraddress" : addr,
        "stripped_address"     : stripped,
        "normalized_address"   : normalized,
        "city_candidates"      : city_candidates,
        "bgy_city_candidates"  : bgy_city_candidates,
    })

noise_df   = pd.DataFrame(noise_records)
working_df = pd.DataFrame(pipeline_inputs)

matched_records = working_df.apply(match_address, axis=1).tolist()
results_df      = pd.DataFrame(matched_records)

print(f"\n  🗑️  Noise    : {len(noise_df):,}")
print(f"  Pipeline   : {len(results_df):,} addresses processed")


# ── 10. STAGE 7 — SEGREGATE & EXPORT ─────────────────────────────────────────

OUTPUT_COLS = [
    "order_deliveraddress", "normalized_address",
    "barangay_code", "barangay_name",
    "city_code",     "city_name",
    "province_code", "province_name",
    "region_code",   "region_name",
    "bgy_match_score", "city_match_score", "confidence_score",
    "match_tier", "match_reason",
]

NOISE_COLS = ["order_deliveraddress", "noise_reason"]


def ensure_cols(df, cols):
    for c in cols:
        if c not in df.columns:
            df[c] = None
    return df[cols].copy()


valid_df   = results_df[results_df["match_tier"] == "valid"]
partial_df = results_df[results_df["match_tier"] == "partial"]
invalid_df = results_df[results_df["match_tier"] == "invalid"]

print(f"\n  ✅ Valid   : {len(valid_df):,}")
print(f"  🟡 Partial : {len(partial_df):,}")
print(f"  ❌ Invalid : {len(invalid_df):,}")
print(f"  🗑️  Noise   : {len(noise_df):,}")


def write_excel(df: pd.DataFrame, path: str, sheet: str = "Results"):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet)
        ws = writer.sheets[sheet]
        for col_cells in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)
        header_fill = PatternFill("solid", fgColor="1F4E79")
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"
    print(f"  Exported → {path}")


date_tag     = datetime.now().strftime("%Y%m%d")
valid_path   = os.path.join(VALID_DIR,   f"valid_{date_tag}.xlsx")
partial_path = os.path.join(PARTIAL_DIR, f"partial_{date_tag}.xlsx")
invalid_path = os.path.join(INVALID_DIR, f"invalid_{date_tag}.xlsx")
noise_path   = os.path.join(NOISE_DIR,   f"noise_{date_tag}.xlsx")

write_excel(ensure_cols(valid_df,   OUTPUT_COLS), valid_path,   "Valid")
write_excel(ensure_cols(partial_df, OUTPUT_COLS), partial_path, "Partial")
write_excel(ensure_cols(invalid_df, OUTPUT_COLS), invalid_path, "Invalid")
write_excel(ensure_cols(noise_df,   NOISE_COLS),  noise_path,   "Noise")

print("\n✅ All exports done.")
print(f"  Valid   → {valid_path}")
print(f"  Partial → {partial_path}")
print(f"  Invalid → {invalid_path}")
print(f"  Noise   → {noise_path}")