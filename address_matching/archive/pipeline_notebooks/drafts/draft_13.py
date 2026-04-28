"""
PH Address Parsing Pipeline  (v2)
==================================
Stage 0 — Load reference data
Stage 1 — Junk token stripping
Stage 2 — Alias normalization
Stage 3 — City detection  (right-to-left, exact → fuzzy, multi-candidate)
Stage 4 — Barangay fuzzy matching  (blocked by city)
Stage 5 — Confidence scoring
Stage 6 — Export  (valid / partial / invalid)
"""

import os
import re
import json
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd
from rapidfuzz import fuzz, process
from openpyxl.styles import Font, PatternFill, Alignment

# ── 0. CONFIG ─────────────────────────────────────────────────────────────────

DIM_LOC_PATH  = "../../data/mapping/dim_location_20260415_v3.csv"
ALIAS_PATH    = "../../data/utils/ph_address_alias_v5.csv"   # augmented — see ph_address_alias_v5.csv
INPUT_PATH    = "../../data/sample_address.xlsx"
INPUT_COL     = "order_deliveraddress"

OUTPUT_DIR    = "../../data/output"
VALID_DIR     = os.path.join(OUTPUT_DIR, "valid")
PARTIAL_DIR   = os.path.join(OUTPUT_DIR, "partial")
INVALID_DIR   = os.path.join(OUTPUT_DIR, "invalid")

for d in [VALID_DIR, PARTIAL_DIR, INVALID_DIR]:
    os.makedirs(d, exist_ok=True)

# Fuzzy thresholds — tune these against your data
CITY_FUZZY_THRESHOLD   = 85   # token_set_ratio: city name match in address segment
BGY_FUZZY_THRESHOLD    = 70   # token_sort_ratio: barangay name match (higher = fewer false positives)
CONFIDENCE_VALID       = 75   # composite score >= this → valid
CONFIDENCE_PARTIAL     = 45   # composite score >= this → partial, else invalid

print("Config OK ✓")


# ── 1. HELPERS ────────────────────────────────────────────────────────────────

def strip_accents(text: str) -> str:
    """Remove diacritical marks (ñ → n, é → e, etc.)."""
    return "".join(
        c for c in unicodedata.normalize("NFD", text)
        if unicodedata.category(c) != "Mn"
    )

def clean_str(s: str) -> str:
    """Lowercase, strip accents, collapse whitespace."""
    s = strip_accents(str(s)).lower()
    return re.sub(r"\s+", " ", s).strip()

print("Helpers defined ✓")


# ── 2. LOAD REFERENCE DATA ────────────────────────────────────────────────────

print("\nLoading reference data …")

# dim_location  (42k+ barangay rows)
dim_raw = pd.read_csv(DIM_LOC_PATH, encoding="latin1")
dim_raw.columns = dim_raw.columns.str.strip()
for col in ["barangay_name", "city_name", "province_name", "region_name"]:
    dim_raw[col] = dim_raw[col].astype(str).str.strip()

dim_raw["bgy_clean"]  = dim_raw["barangay_name"].apply(clean_str)
dim_raw["city_clean"] = dim_raw["city_name"].apply(clean_str)
dim_raw["prov_clean"] = dim_raw["province_name"].apply(clean_str)

# ── OPT 1: Pre-sort cities once — never sort inside the per-address loop ──────
all_cities_clean  = {clean_str(c): c for c in dim_raw["city_name"].dropna().unique()}
_sorted_cities    = sorted(all_cities_clean.items(), key=lambda x: -len(x[0]))
all_city_cleans   = [c for c, _ in _sorted_cities]   # for rapidfuzz list

# ── OPT 2: Compile one mega alternation regex — replaces 1407-iteration loop ──
# Sorted longest-first so the regex engine matches the most specific name first.
# e.g. "San Fernando" matches before "San" would.
_city_alts   = "|".join(re.escape(city_c) for city_c, _ in _sorted_cities)
CITY_MEGA_RE = re.compile(r"\b(?:" + _city_alts + r")\b", re.IGNORECASE)
print(f"  City mega-regex compiled ({len(_sorted_cities)} cities) ✓")

# ── OPT 3a: Precompute city subsets dict — O(1) lookup replaces df filter ─────
# city_clean → DataFrame subset for that city (used in barangay matching)
_city_subsets: dict[str, pd.DataFrame] = {
    city_c: grp.copy()
    for city_c, grp in dim_raw.groupby("city_clean")
}
print(f"  City subsets precomputed ({len(_city_subsets)} keys) ✓")

# ── OPT 3b: Precompute barangay exact-match dict — O(1) replaces df filter ───
# bgy_clean → [(city_name, province_name), ...]
_bgy_exact_dict: dict[str, list[tuple[str, str]]] = {}
for _, row in dim_raw.iterrows():
    _bgy_exact_dict.setdefault(row["bgy_clean"], []).append(
        (row["city_name"], row["province_name"])
    )
print(f"  Barangay exact dict precomputed ({len(_bgy_exact_dict):,} keys) ✓")

# ── OPT 3c: Precompute stripped barangay list for fuzzy fallback ──────────────
# Stripping "barangay" prefix avoids false 100-score matches on "barangay X"
_BARANGAY_PREFIX_RE = re.compile(r"^\s*barangay\s*", re.IGNORECASE)
_bgy_stripped_entries: list[tuple[str, str, str, str]] = [
    (row["bgy_clean"],
     _BARANGAY_PREFIX_RE.sub("", row["bgy_clean"]).strip(),
     row["city_name"],
     row["province_name"])
    for _, row in dim_raw.iterrows()
]
_bgy_stripped_list = [x[1] for x in _bgy_stripped_entries]
print(f"  Barangay stripped list precomputed ({len(_bgy_stripped_list):,} entries) ✓")

# Alias map  (abbreviation/shorthand → canonical form)
alias_df = pd.read_csv(ALIAS_PATH, encoding="latin1", usecols=["pattern", "replacement"])
alias_df = alias_df.dropna(subset=["pattern", "replacement"])
alias_df["pattern"]     = alias_df["pattern"].astype(str).str.strip()
alias_df["replacement"] = alias_df["replacement"].astype(str).str.strip()
alias_df = alias_df.sort_values("pattern", key=lambda s: s.str.len(), ascending=False)
alias_map = dict(zip(alias_df["pattern"].str.upper(), alias_df["replacement"].str.upper()))

print(f"  dim_location rows : {len(dim_raw):,}")
print(f"  alias rules       : {len(alias_map):,}")
print(f"  unique cities     : {len(all_cities_clean):,}")
print("Reference data loaded ✓")


# ── 3. LOAD ADDRESSES ─────────────────────────────────────────────────────────

print("\nLoading addresses …")
input_df      = pd.read_excel(INPUT_PATH)
RAW_ADDRESSES = input_df[INPUT_COL].dropna().tolist()
print(f"  {len(RAW_ADDRESSES):,} addresses loaded")


# ── 4. STAGE 1 — JUNK TOKEN STRIPPING ────────────────────────────────────────
#
# Removes noise that hurts fuzzy matching:
#   • Stray punctuation (backticks, tildes)
#   • Phone numbers
#   • Landmark phrases  (near, beside, in front of, across, opp.)
#   • Lot / Block / Unit / Floor / Room designations
#   • Building / subdivision / compound names — unified single pass that strips
#     0-4 PREFIX words + the keyword + 0-4 SUFFIX words in one go.
#     This prevents the prefix being left orphaned when the keyword is removed first.
#   • Street-level numbers (3-5 digits) — kept: 2-digit brgy numbers
#
# Meaningful abbreviations are protected before stripping and restored after.
# e.g.  A.C → ANGELES_CITY → restored after junk removal

_BUILDING_KW = (
    r"bldg|building|tower|plaza|centre|center|subd|subdivision|compound|cmpd|"
    r"village|vill|estate|residences|residencia|condominium|condo|"
    r"apartelle|apartment|apt|annex|mall|square|complex|commercial|industrial|zone|cluster"
)

_PROTECT_PATTERNS = [
    # A.C / A.C. = Angeles City — keep it through stripping so alias map can expand it
    (re.compile(r"\bA\.C\.?\b", re.IGNORECASE), "ANGELES_CITY"),
]

_STRAY_PUNCT      = re.compile(r"[`~|]")
_PHONE_NUMBERS    = re.compile(r"\b(0|\+63)\d{9,10}\b")
_STREET_NUMBERS   = re.compile(r"(?<!\w)\d{3,5}(?!\w)")
_LOT_BLK_UNIT     = re.compile(
    r"\b(lot|blk|block|unit|floor|flr|fl|rm|room|door|phase|house no|hse no|#)"
    r"\s*[.\-]?\s*[\w\-]*",
    re.IGNORECASE,
)
_LANDMARK_PHRASES = re.compile(
    r"\b(near|beside|in front of|across from|across|opposite|opp\.?|behind|"
    r"adjacent to|adj\.?|along|corner of|corner|cor\.?)\b[^,]*",
    re.IGNORECASE,
)

# Unified building pattern: [0-4 prefix words] [keyword] [0-4 suffix words]
# Suffix stops at known address anchors (Brgy, St, Ave, Road) to prevent over-stripping.
_BUILDING_UNIFIED = re.compile(
    r"\b(?:[A-Za-z]\w*\.?\s+){0,4}"
    r"\b(?:" + _BUILDING_KW + r")\b\.?"
    r"(?:\s+(?!(?:brgy|barangay|st\b|ave\b|road\b|blvd\b))[A-Za-z]\w*){0,4}",
    re.IGNORECASE,
)


def strip_junk(addr: str) -> str:
    """
    Remove noise tokens from a raw address string.
    Returns a cleaner string ready for alias normalization.
    Extend _BUILDING_KW or _PROTECT_PATTERNS as new noise patterns emerge.
    """
    s = addr
    # Protect meaningful abbreviations
    for pattern, placeholder in _PROTECT_PATTERNS:
        s = pattern.sub(placeholder, s)

    s = _STRAY_PUNCT.sub(" ", s)
    s = _PHONE_NUMBERS.sub(" ", s)
    s = _LANDMARK_PHRASES.sub(" ", s)
    s = _LOT_BLK_UNIT.sub(" ", s)
    s = _BUILDING_UNIFIED.sub(" ", s)   # single unified pass — prefix+keyword+suffix
    s = _STREET_NUMBERS.sub(" ", s)

    # Restore protected tokens
    s = s.replace("ANGELES_CITY", "Angeles City")

    s = re.sub(r"\s{2,}", " ", s)
    s = re.sub(r"(,\s*){2,}", ", ", s)
    s = re.sub(r"^[\s,\.]+|[\s,\.]+$", "", s)
    return s.strip()


print("Junk stripper defined ✓")


# ── 5. STAGE 2 — ALIAS NORMALIZATION ─────────────────────────────────────────
#
# Applies alias_map token-by-token (up to 3-word windows, longest-first).
# Carried over from the current pipeline — this part works well.

def normalize_address(addr: str) -> str:
    """
    Apply alias replacements on up to 3-word windows, longest-first.
    Returns title-cased normalized string.
    """
    upper  = addr.upper()
    tokens = re.split(r"(\s+|,|\.)", upper)
    result = []
    i = 0
    while i < len(tokens):
        tok = tokens[i].strip()
        if not tok or tok in (",", "."):
            result.append(tokens[i])
            i += 1
            continue
        matched = False
        for lookahead in (3, 2, 1):
            candidate_tokens, j, count = [], i, 0
            while j < len(tokens) and count < lookahead:
                if re.match(r"^\s*$", tokens[j]) or tokens[j] in (",", "."):
                    j += 1
                    continue
                candidate_tokens.append(tokens[j])
                j += 1
                count += 1
            candidate = " ".join(candidate_tokens)
            if candidate in alias_map:
                result.append(alias_map[candidate])
                i = j
                matched = True
                break
        if not matched:
            result.append(tokens[i])
            i += 1
    normalized = re.sub(r"\s+", " ", "".join(result)).strip()
    # Remove stray dots left when "BRGY." is tokenized as BRGY + "."
    normalized = re.sub(r"\.\s+", " ", normalized)
    normalized = re.sub(r"\s{2,}", " ", normalized).strip()
    return normalized.title()


print("Alias normalizer defined ✓")


# ── 6. STAGE 3 — CITY DETECTION ──────────────────────────────────────────────
#
# Two functions work together:
#
# detect_city_candidates()  — primary: scans right-to-left across comma segments.
#   Per segment, 4 passes:
#     Pass 1 — CITY_MEGA_RE single scan (replaces 1407-iteration loop, ~270x faster)
#     Pass 2 — Strip trailing "City" from segment and re-scan with CITY_MEGA_RE
#              (handles alias-expanded "Angeles City" vs dim "Angeles",
#               generalises to Davao City, Cebu City, Iloilo City, etc.)
#     Pass 3 — rapidfuzz token_set_ratio fallback on original segment (≥ CITY_FUZZY_THRESHOLD)
#     Pass 4 — rapidfuzz fallback on City-stripped segment
#
# detect_via_barangay() — fallback when city detection yields nothing.
#   Strips "Barangay" prefix + number tokens + street type words, then:
#     Step 1 — O(1) dict lookup against _bgy_exact_dict (3→2→1 word windows)
#     Step 2 — rapidfuzz token_sort_ratio over _bgy_stripped_list (≥ BGY_FUZZY_THRESHOLD)
#   Returns [(city_original, province_original), ...] inferred from barangay name.

_STREET_TOKENS_RE = re.compile(
    r"\b(street|avenue|road|boulevard|highway|lane|drive|circle|place|extension)\b",
    re.IGNORECASE,
)


def _prep_for_bgy_match(norm_addr: str) -> str:
    """Strip barangay prefix, number tokens, and street types before bgy fuzzy match."""
    s = _BARANGAY_PREFIX_RE.sub("", clean_str(norm_addr)).strip()
    s = re.sub(r"\b[\d][\d\-a-z]*\b", "", s)   # e.g. 22, 12-A, 54
    s = _STREET_TOKENS_RE.sub("", s)
    return re.sub(r"\s{2,}", " ", s).strip()


def detect_city_candidates(norm_addr: str) -> list[tuple[str, str]]:
    """
    Returns [(city_clean, city_original), ...] detected in the address.
    Empty list → fall through to detect_via_barangay().

    OPT: Uses CITY_MEGA_RE (single compiled alternation) instead of a
         per-city loop — ~270x faster on cache misses.
    """
    segments_rtl = list(reversed([s.strip() for s in norm_addr.split(",")]))
    candidates: dict[str, str] = {}

    for seg in segments_rtl:
        seg_c = clean_str(seg)
        if len(seg_c) < 3:
            continue

        # Pass 1 — mega regex scan (single pass over segment)
        m = CITY_MEGA_RE.search(seg_c)
        if m:
            matched = m.group(0).lower()
            candidates[matched] = all_cities_clean[matched]
            continue

        # Pass 2 — strip "City" suffix from segment tokens and re-scan
        seg_nc = re.sub(r"\s{2,}", " ", re.sub(r"\bcity\b", "", seg_c)).strip()
        if seg_nc != seg_c:
            m2 = CITY_MEGA_RE.search(seg_nc)
            if m2:
                matched = m2.group(0).lower()
                candidates[matched] = all_cities_clean[matched]
                continue

        # Pass 3 — fuzzy fallback on original segment
        best = process.extractOne(seg_c, all_city_cleans, scorer=fuzz.token_set_ratio)
        if best and best[1] >= CITY_FUZZY_THRESHOLD:
            candidates[best[0]] = all_cities_clean[best[0]]
            continue

        # Pass 4 — fuzzy on City-stripped segment
        if seg_nc != seg_c:
            best2 = process.extractOne(seg_nc, all_city_cleans, scorer=fuzz.token_set_ratio)
            if best2 and best2[1] >= CITY_FUZZY_THRESHOLD:
                candidates[best2[0]] = all_cities_clean[best2[0]]

    return list(candidates.items())


def detect_via_barangay(norm_addr: str) -> list[tuple[str, str]]:
    """
    Fallback: infer city from barangay name when city detection yields nothing.
    Returns [(city_original, province_original), ...].

    OPT: Uses _bgy_exact_dict for O(1) exact n-gram lookup instead of
         repeated df[df["bgy_clean"] == phrase] filter calls.
    """
    q = _prep_for_bgy_match(norm_addr)
    if len(q) < 3:
        return []

    # Step 1 — O(1) exact n-gram lookup (3 → 2 → 1 word windows)
    tokens = q.split()
    for n in (3, 2, 1):
        for i in range(len(tokens) - n + 1):
            phrase = " ".join(tokens[i:i + n])
            if len(phrase) < 4:
                continue
            if phrase in _bgy_exact_dict:
                seen: dict[str, str] = {}
                for city, prov in _bgy_exact_dict[phrase]:
                    seen.setdefault(city, prov)
                return list(seen.items())

    # Step 2 — fuzzy fallback over stripped barangay list
    best = process.extractOne(
        q, _bgy_stripped_list,
        scorer=fuzz.token_sort_ratio,
        score_cutoff=BGY_FUZZY_THRESHOLD,   # OPT: exit early, skips weak candidates
    )
    if best:
        idx   = _bgy_stripped_list.index(best[0])
        entry = _bgy_stripped_entries[idx]
        return [(entry[2], entry[3])]

    return []


print("City detector + barangay fallback defined ✓")


# ── 7. STAGE 4 — BARANGAY FUZZY MATCHING ─────────────────────────────────────
#
# For each (city_clean, city_original) candidate:
#   • Sub-filter dim_location to that city
#   • Fuzzy-match the cleaned address against bgy_clean  (partial_ratio)
#   • Collect best (score, row) per candidate
# Then pick the overall best-scoring candidate.
#
# Returns a result dict that includes the match, its score, and a match_tier.

def _score_match(addr_c: str, city_c: str) -> tuple[float, pd.Series | None]:
    """
    Fuzzy-match addr_c against barangay names for city_c.
    Uses precomputed _city_subsets dict — O(1) lookup, no DataFrame filter.
    Returns (best_score, best_row).  score = 0.0 if no match found.
    """
    sub = _city_subsets.get(city_c)
    if sub is None or sub.empty:
        return 0.0, None
    bgy_names   = sub["bgy_clean"].tolist()
    best        = process.extractOne(addr_c, bgy_names, scorer=fuzz.partial_ratio)
    if best is None:
        return 0.0, None
    matched_row = sub[sub["bgy_clean"] == best[0]].iloc[0]
    return float(best[1]), matched_row


def match_address(row: pd.Series) -> dict:
    """
    Full matching for one address row.
    Returns a flat dict of all output columns.
    """
    norm           = row["normalized_address"]
    addr_c         = clean_str(norm)
    candidates     = list(row["city_candidates"])       # [(city_clean, city_original), ...]
    bgy_city_cands = row["bgy_city_candidates"]         # [(city_orig, prov_orig), ...] or []

    # ── Promote barangay fallback into candidates when city detect found nothing
    if not candidates and bgy_city_cands:
        candidates = [(clean_str(city), city) for city, prov in bgy_city_cands]

    # ── Nothing at all ────────────────────────────────────────────────────────
    if not candidates:
        return _build_result(
            orig_addr  = row["order_deliveraddress"],
            norm_addr  = norm,
            addr_c     = addr_c,
            matched_row= None,
            bgy_score  = 0.0,
            city_score = 0.0,
            tier       = "invalid",
            reason     = "no_city_detected",
        )

    # ── Try each city candidate, keep best barangay score ────────────────────
    best_bgy_score  = -1.0
    best_row        = None
    best_city_clean = None
    best_city_orig  = None

    for city_c, city_orig in candidates:
        bgy_score, matched_row = _score_match(addr_c, city_c)   # OPT: dict lookup, no df filter
        if bgy_score > best_bgy_score:
            best_bgy_score  = bgy_score
            best_row        = matched_row
            best_city_clean = city_c
            best_city_orig  = city_orig

    # City-presence score: is the detected city string literally in the address?
    city_score = fuzz.partial_ratio(best_city_clean, addr_c) if best_city_clean else 0.0

    # Composite confidence  (70% barangay match, 30% city presence)
    composite = 0.7 * best_bgy_score + 0.3 * city_score

    # ── Determine tier ────────────────────────────────────────────────────────
    if best_bgy_score >= BGY_FUZZY_THRESHOLD and composite >= CONFIDENCE_VALID:
        tier   = "valid"
        reason = "barangay_matched"
    elif composite >= CONFIDENCE_PARTIAL:
        tier   = "partial"
        reason = "city_only_or_low_confidence"
    else:
        tier   = "invalid"
        reason = "below_confidence_threshold"

    return _build_result(
        orig_addr   = row["order_deliveraddress"],
        norm_addr   = norm,
        addr_c      = addr_c,
        matched_row = best_row,
        bgy_score   = round(best_bgy_score, 1),
        city_score  = round(city_score, 1),
        tier        = tier,
        reason      = reason,
        city_orig   = best_city_orig,
        composite   = round(composite, 1),
    )


def _build_result(
    orig_addr, norm_addr, addr_c,
    matched_row, bgy_score, city_score,
    tier, reason,
    city_orig=None, composite=None,
) -> dict:
    """Assemble the output dict from a matched (or unmatched) row."""
    if matched_row is not None:
        return {
            "order_deliveraddress" : orig_addr,
            "normalized_address"   : norm_addr,
            "barangay_code"        : matched_row.get("barangay_code"),
            "barangay_name"        : matched_row["barangay_name"],
            "city_code"            : matched_row.get("city_code"),
            "city_name"            : matched_row["city_name"],
            "province_code"        : matched_row.get("province_code"),
            "province_name"        : matched_row["province_name"],
            "region_code"          : matched_row.get("region_code"),
            "region_name"          : matched_row["region_name"],
            "addr_clean"           : addr_c,
            "bgy_match_score"      : bgy_score,
            "city_match_score"     : city_score,
            "confidence_score"     : composite,
            "match_tier"           : tier,
            "match_reason"         : reason,
        }
    else:
        return {
            "order_deliveraddress" : orig_addr,
            "normalized_address"   : norm_addr,
            "barangay_code"        : None,
            "barangay_name"        : None,
            "city_code"            : None,
            "city_name"            : city_orig,
            "province_code"        : None,
            "province_name"        : None,
            "region_code"          : None,
            "region_name"          : None,
            "addr_clean"           : addr_c,
            "bgy_match_score"      : bgy_score,
            "city_match_score"     : city_score,
            "confidence_score"     : composite if composite is not None else 0.0,
            "match_tier"           : tier,
            "match_reason"         : reason,
        }


print("Matcher defined ✓")


# ── 8. RUN PIPELINE ───────────────────────────────────────────────────────────
#
# Single-threaded loop with tqdm progress bar.
# For 120k rows (~8 min single-thread), parallelise with:
#
#   from multiprocessing import Pool
#   with Pool(processes=os.cpu_count()) as pool:
#       results = pool.map(process_one, RAW_ADDRESSES)
#
# where process_one() wraps stages 1-5 for a single address.
# See the parallel runner template at the bottom of this file.

try:
    from tqdm import tqdm
    _iter = tqdm(RAW_ADDRESSES, desc="Parsing", unit="addr")
except ImportError:
    _iter = RAW_ADDRESSES
    print("  (install tqdm for a progress bar: pip install tqdm)")

print("\nRunning pipeline …")

records = []
for addr in _iter:
    # Stage 1 — strip junk
    stripped = strip_junk(addr)
    # Stage 2 — normalize aliases
    normalized = normalize_address(stripped)
    # Stage 3a — detect city from address directly
    city_candidates = detect_city_candidates(normalized)
    # Stage 3b — barangay fallback (only computed when city detect fails)
    bgy_city_candidates = detect_via_barangay(normalized) if not city_candidates else []
    records.append({
        "order_deliveraddress" : addr,
        "stripped_address"     : stripped,
        "normalized_address"   : normalized,
        "city_candidates"      : city_candidates,
        "bgy_city_candidates"  : bgy_city_candidates,
    })

working_df = pd.DataFrame(records)

# Stage 4+5 — match + score
matched_records = working_df.apply(match_address, axis=1).tolist()
results_df      = pd.DataFrame(matched_records)

print(f"Pipeline complete — {len(results_df):,} addresses processed")


# ── 9. STAGE 6 — SEGREGATE & EXPORT ──────────────────────────────────────────

OUTPUT_COLS = [
    "order_deliveraddress", "normalized_address",
    "barangay_code", "barangay_name",
    "city_code",     "city_name",
    "province_code", "province_name",
    "region_code",   "region_name",
    "bgy_match_score", "city_match_score", "confidence_score",
    "match_tier", "match_reason",
]

def ensure_cols(df, cols):
    for c in cols:
        if c not in df.columns:
            df[c] = None
    return df[cols].copy()

valid_df   = results_df[results_df["match_tier"] == "valid"]
partial_df = results_df[results_df["match_tier"] == "partial"]
invalid_df = results_df[results_df["match_tier"] == "invalid"]

print(f"\n  ✅ Valid   : {len(valid_df):,}")
print(f"  🟡 Partial : {len(partial_df):,}")
print(f"  ❌ Invalid : {len(invalid_df):,}")


def write_excel(df: pd.DataFrame, path: str, sheet: str = "Results"):
    """Write a styled Excel file with frozen header row."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet)
        ws = writer.sheets[sheet]
        for col_cells in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)
        header_fill = PatternFill("solid", fgColor="1F4E79")
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"
    print(f"  Exported → {path}")


date_tag     = datetime.now().strftime("%Y%m%d")
valid_path   = os.path.join(VALID_DIR,   f"valid_{date_tag}.xlsx")
partial_path = os.path.join(PARTIAL_DIR, f"partial_{date_tag}.xlsx")
invalid_path = os.path.join(INVALID_DIR, f"invalid_{date_tag}.xlsx")

write_excel(ensure_cols(valid_df,   OUTPUT_COLS), valid_path,   "Valid")
write_excel(ensure_cols(partial_df, OUTPUT_COLS), partial_path, "Partial")
write_excel(ensure_cols(invalid_df, OUTPUT_COLS), invalid_path, "Invalid")

print("\n✅ All exports done.")
print(f"  Valid   → {valid_path}")
print(f"  Partial → {partial_path}")
print(f"  Invalid → {invalid_path}")


# ══════════════════════════════════════════════════════════════════════════════
# OPTIONAL: PARALLEL RUNNER  (~4x faster on 4 cores, ~8x on 8 cores)
# ══════════════════════════════════════════════════════════════════════════════
#
# Replace the run loop above with this block for large batches (50k+ rows).
# All precomputed structures (CITY_MEGA_RE, _city_subsets, _bgy_exact_dict,
# _bgy_stripped_list) are module-level globals and are inherited by worker
# processes via fork (Linux/Mac) without re-serialisation overhead.
#
# Usage: uncomment, set N_WORKERS, replace the run loop section above.
#
# import os
# from multiprocessing import Pool
#
# N_WORKERS = os.cpu_count()   # or set explicitly, e.g. 8
#
# def _process_one(addr: str) -> dict:
#     """Stages 1-5 for a single address — safe to call from worker processes."""
#     stripped            = strip_junk(addr)
#     normalized          = normalize_address(stripped)
#     city_candidates     = detect_city_candidates(normalized)
#     bgy_city_candidates = detect_via_barangay(normalized) if not city_candidates else []
#     row = pd.Series({
#         "order_deliveraddress" : addr,
#         "stripped_address"     : stripped,
#         "normalized_address"   : normalized,
#         "city_candidates"      : city_candidates,
#         "bgy_city_candidates"  : bgy_city_candidates,
#     })
#     return match_address(row)
#
# if __name__ == "__main__":   # required guard for multiprocessing on Windows
#     with Pool(processes=N_WORKERS) as pool:
#         matched_records = list(tqdm(
#             pool.imap(_process_one, RAW_ADDRESSES, chunksize=500),
#             total=len(RAW_ADDRESSES), desc="Parsing (parallel)", unit="addr"
#         ))
#     results_df = pd.DataFrame(matched_records)

